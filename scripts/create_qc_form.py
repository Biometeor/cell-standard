from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

wb = Workbook()

# Colors
HEADER_FILL = PatternFill('solid', fgColor='4472C4')
HEADER_FONT = Font(name='Microsoft YaHei', bold=True, size=10, color='FFFFFF')
SUB_FILL = PatternFill('solid', fgColor='D6E4F0')
SUB_FONT = Font(name='Microsoft YaHei', bold=True, size=10, color='1F4E79')
TITLE_FONT = Font(name='Microsoft YaHei', bold=True, size=14, color='1F4E79')
SECTION_FONT = Font(name='Microsoft YaHei', bold=True, size=11, color='2E75B6')
NORMAL = Font(name='Microsoft YaHei', size=10)
BOLD = Font(name='Microsoft YaHei', bold=True, size=10)
WHITE_FILL = PatternFill('solid', fgColor='FFFFFF')
INPUT_FILL = PatternFill('solid', fgColor='F0F8FF')
LEVEL1_FILL = PatternFill('solid', fgColor='C6EFCE')
LEVEL2_FILL = PatternFill('solid', fgColor='FFEB9C')
LEVEL3_FILL = PatternFill('solid', fgColor='FFC7CE')
MUST_FILL = PatternFill('solid', fgColor='FFF2CC')
NOTE_FONT = Font(name='Microsoft YaHei', size=9, italic=True, color='666666')
BORDER = Border(
    left=Side(style='thin', color='BFBFBF'),
    right=Side(style='thin', color='BFBFBF'),
    top=Side(style='thin', color='BFBFBF'),
    bottom=Side(style='thin', color='BFBFBF')
)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
WRAP = Alignment(wrap_text=True, vertical='center')

def hdr(ws, row, ncols):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER

def dc(ws, row, col, val, font=None, fill=None, align=None):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = font or NORMAL
    cell.alignment = align or WRAP
    cell.border = BORDER
    if fill:
        cell.fill = fill

# ============================================================
# Sheet 1: 阈值参考
# ============================================================
ws = wb.active
ws.title = '阈值参考'

ws.merge_cells('A1:G1')
ws['A1'] = 'QC 指标阈值参考（填写前请阅读）'
ws['A1'].font = TITLE_FONT

r = 3
ws.merge_cells(f'A{r}:G{r}')
ws[f'A{r}'] = '填写时打开对应类型的 Sheet，每行填一个样本，指标值填入对应列中'
ws[f'A{r}'].font = NOTE_FONT

# --- scRNA-seq reference ---
r = 5
ws.merge_cells(f'A{r}:G{r}')
ws[f'A{r}'] = '▼ scRNA-seq / snRNA-seq'
ws[f'A{r}'].font = SECTION_FONT
r += 1

ref_headers = ['指标', '必填/可选', '1级（优质）', '2级（合格）', '3级（让步）', '定义', '单位']
for i, h in enumerate(ref_headers, 1):
    ws.cell(row=r, column=i, value=h)
hdr(ws, r, 7)
r += 1

scrna_ref = [
    ['total_valid_cells', '必填', '> 2000', '1000-2000', '< 1000', '过滤后有效细胞数', '个'],
    ['Median Genes per Cell', '必填', '> 2000', '1000-2000', '500-1000', '每个细胞基因数中位数', '个'],
    ['Median UMI per Cell', '必填', '> 5000', '1000-5000', '500-1000', '每个细胞UMI数中位数', '个'],
    ['Low Quality Cell Rate', '必填', '< 10%', '10%-25%', '25%-40%', '低基因数(<500)细胞占比', '%'],
    ['Mitochondrial Fraction', '参考', '—', '—', '—', '线粒体UMI占比，通常<10%', '%'],
    ['Ribosomal Fraction', '参考', '—', '—', '—', '核糖体UMI占比', '%'],
    ['sparsity', '参考', '—', '—', '—', '矩阵零元占比', '%'],
    ['Doublet Rate', '可选', '< 5%', '5%-10%', '10%-20%', '预测双细胞比例', '%'],
    ['Cell Cluster Separation', '可选', '> 0.25', '0.1-0.25', '< 0.1', '轮廓系数聚类分离度', ''],
    ['Batch Effect Removal R²', '可选', '< 0.05', '0.05-0.15', '> 0.15', '批次因素解释方差(越小越好)', ''],
    ['Cell Type Annotation Rate', '可选', '> 90%', '70%-90%', '< 70%', '成功注释细胞比例', '%'],
    ['Cell Type Consistency', '可选', '> 90%', '70%-90%', '< 70%', '标签在样本间一致性', '%'],
]

for row_data in scrna_ref:
    for c, val in enumerate(row_data, 1):
        fill = MUST_FILL if c == 2 and val == '必填' else None
        dc(ws, r, c, val, fill=fill)
    r += 1

# --- Spatial reference ---
r += 1
ws.merge_cells(f'A{r}:G{r}')
ws[f'A{r}'] = '▼ Spatial Transcriptomics'
ws[f'A{r}'].font = SECTION_FONT
r += 1

for i, h in enumerate(ref_headers[:6], 1):
    ws.cell(row=r, column=i, value=h)
hdr(ws, r, 6)
r += 1

spatial_ref = [
    ['spot_count', '必填', '≥ 2000', '个', '组织范围内有UMI的有效spot数'],
    ['gene_count', '必填', '≥ 5000', '个', '检测到的unique gene种类数'],
    ['median_genes_per_spot', '必填', '≥ 500', '个', '每个spot中位基因数'],
    ['mean_genes_per_spot', '必填', '≥ 400', '个', '每个spot平均基因数'],
    ['median_counts_per_spot', '必填', '≥ 1000', '个', '每个spot中位计数'],
    ['mean_umi_per_spot', '必填', '≥ 800', '个', '每个spot平均UMI数'],
    ['Unique_Reads', '可选', '≥ 100M', '个', '纠错后去重reads数(Stereo-seq)'],
    ['Total_Reads', '可选', '≥ 1G', '个', '总reads数'],
    ['Valid_Barcode_Reads', '可选', '≥ 75%', '%', '有效分子标签reads比例(Stereo-seq)'],
    ['Unique_Mapped_Reads', '可选', '≥ 70%', '%', '唯一比对参考基因组的reads比例(Stereo-seq)'],
    ['rna_mapping', '可选', '≥ 85%', '%', '整体RNA比对率'],
    ['mean_umi_per_Bin200', '可选', '≥ 50000', '个', 'Bin200下平均UMI数(Stereo-seq)'],
    ['Mean Gene Type per Bin200', '可选', '≥ 5000', '个', 'Bin200下平均基因种类数(Stereo-seq)'],
    ['tissue_coverage', '可选', '≥ 50%', '%', '组织覆盖率'],
    ['fraction_of_spots_under_tissue', '可选', '≥ 70%', '%', '组织下spot比例'],
    ['tissue_area_coverage', '可选', '≥ 60%', '%', '组织面积覆盖率'],
    ['low_information_spot_rate', '可选', '≤ 10%', '%', '低信息spot比例(基因<50或UMI<100)'],
    ['image_resolution', '可选', '≥ 1024×1024', 'px', '图像分辨率'],
    ['image_contrast_PSNR', '可选', '≥ 20', 'dB', '图像对比度'],
    ['alignment_quality', '可选', '≥ 0.8', '评分', '图像与spot对齐准确度'],
    ['cluster_marker_gene_specificity', '可选', '≥ 70%', '%', 'marker基因在目标聚类区域表达占比'],
    ['tissue_region_gene_expression_consistency', '可选', '≥ 60%', '%', '已知组织区域特征基因表达匹配度'],
    ['mitochondrial_percentage', '参考', '统计参考', '%', '线粒体基因比例(植物中重要)'],
    ['ribosomal_percentage', '参考', '统计参考', '%', '核糖体基因比例'],
]

for row_data in spatial_ref:
    for c, val in enumerate(row_data, 1):
        fill = MUST_FILL if c == 2 and val == '必填' else None
        dc(ws, r, c, val, fill=fill)
    r += 1

# --- scATAC-seq reference ---
r += 1
ws.merge_cells(f'A{r}:G{r}')
ws[f'A{r}'] = '▼ scATAC-seq'
ws[f'A{r}'].font = SECTION_FONT
r += 1

for i, h in enumerate(ref_headers, 1):
    ws.cell(row=r, column=i, value=h)
hdr(ws, r, 7)
r += 1

atac_ref = [
    ['Estimated Number of Cells', '必填', '> 预期70%', '预期40%-70%', '< 预期40%', '有效细胞核数量(与预期比率)', ''],
    ['Median Fragments per Cell', '必填', '> 10000', '3000-10000', '1500-3000', '有效unique fragments中位数', '个'],
    ['TSS Enrichment Score', '必填', '> 4', '1-3', '1', 'TSS附近fragment富集(核心)', ''],
    ['FRiP', '必填', '> 50%', '30%-50%', '20%-30%', '落在Peak区域的reads占比', '%'],
    ['Nucleosome Signal', '必填', '< 1.5', '1.5-2.0', '> 2.0', '核小体信号强度(越低越好)', ''],
    ['Doublet Rate', '必填', '< 5%', '5%-10%', '10%-20%', '预测双细胞核比例', '%'],
    ['Low Quality Cell Rate', '必填', '< 10%', '10%-25%', '25%-40%', '低质量细胞核比例', '%'],
    ['Capture Efficiency', '可选', '> 70%', '50%-70%', '< 50%', '有效细胞核捕获率', '%'],
    ['Cell Cluster Separation', '可选', '> 0.2', '0.1-0.2', '< 0.1', '聚类轮廓系数', ''],
    ['Batch Effect Removal R²', '可选', '< 0.05', '0.05-0.15', '> 0.15', '批次效应解释方差(越小越好)', ''],
    ['Cell Type Annotation Rate', '可选', '> 90%', '70%-90%', '< 70%', '基于基因活性/Peak注释比例', '%'],
    ['Cell Type Consistency', '可选', '> 90%', '70%-90%', '< 70%', '标签在样本间一致性', '%'],
]

for row_data in atac_ref:
    for c, val in enumerate(row_data, 1):
        fill = MUST_FILL if c == 2 and val == '必填' else None
        dc(ws, r, c, val, fill=fill)
    r += 1

ws.column_dimensions['A'].width = 36
ws.column_dimensions['B'].width = 10
ws.column_dimensions['C'].width = 16
ws.column_dimensions['D'].width = 16
ws.column_dimensions['E'].width = 16
ws.column_dimensions['F'].width = 35
ws.column_dimensions['G'].width = 10

# ============================================================
# Helper: create data sheet with metrics as columns
# ============================================================
def create_data_sheet(wb, title, metrics, ref_cols_needed=True):
    ws = wb.create_sheet(title)
    ws.title = title

    # Row 1: Title
    ws.merge_cells('A1:ZZ1')
    ws['A1'] = f'{title} — 每行一个样本，绿色列为必填指标'
    ws['A1'].font = TITLE_FONT
    ws.row_dimensions[1].height = 30

    # Row 2: empty for spacing

    # Row 3: Column headers — metric names
    cols = ['样本名'] + [m[0] for m in metrics] + ['备注']
    ncols = len(cols)
    for c, name in enumerate(cols, 1):
        dc(ws, 3, c, name, HEADER_FONT, HEADER_FILL, CENTER)

    # Row 4: Type row — 必填/可选
    types = [''] + [m[1] for m in metrics] + ['']
    for c, t in enumerate(types, 1):
        fill = MUST_FILL if t == '必填' else (SUB_FILL if t else None)
        dc(ws, 4, c, t, SUB_FONT, fill, CENTER)
    ws.row_dimensions[4].height = 22

    # Row 5: Unit row
    units = [''] + [m[-1] for m in metrics] + ['']
    for c, u in enumerate(units, 1):
        dc(ws, 5, c, f'({u})' if u else '', NORMAL, None, CENTER)

    # Data rows (20 rows for filling)
    for data_row in range(6, 26):
        for c in range(1, ncols+1):
            if c == 1:
                dc(ws, data_row, c, '', NORMAL, None, CENTER)
            elif c == ncols:
                dc(ws, data_row, c, '', NORMAL)
            else:
                # Check if this metric is required
                idx = c - 2
                is_req = idx < len(metrics) and metrics[idx][1] == '必填'
                dc(ws, data_row, c, '', NORMAL, INPUT_FILL if is_req else WHITE_FILL)

    # Set column widths
    ws.column_dimensions['A'].width = 18
    for c in range(2, ncols):
        ws.column_dimensions[chr(64+c) if c <= 26 else 'A'].width = 18
    # More robust column width setting
    for i, (name, *_) in enumerate(metrics):
        col_letter = chr(66+i) if i < 25 else 'AA'  # B=66
        try:
            ws.column_dimensions[col_letter].width = max(15, min(22, len(name)*1.2 + 4))
        except:
            ws.column_dimensions[col_letter].width = 18
    last_col = chr(64+ncols) if ncols <= 26 else 'ZZ'
    try:
        ws.column_dimensions[last_col].width = 20
    except:
        pass

    ws.auto_filter.ref = f'A3:{chr(64+ncols)}{25}'

    return ws

# ============================================================
# Metrics definitions
# ============================================================
scrna_req = [
    ('total_valid_cells', '必填', '个'),
    ('Median Genes per Cell', '必填', '个'),
    ('Median UMI per Cell', '必填', '个'),
    ('Low Quality Cell Rate', '必填', '%'),
]
scrna_refs = [
    ('Mitochondrial Fraction', '参考', '%'),
    ('Ribosomal Fraction', '参考', '%'),
    ('sparsity', '参考', '%'),
]
scrna_opt = [
    ('Doublet Rate', '可选', '%'),
    ('Cell Cluster Separation', '可选', ''),
    ('Batch Effect Removal R²', '可选', ''),
    ('Cell Type Annotation Rate', '可选', '%'),
    ('Cell Type Consistency', '可选', '%'),
]

spatial_req = [
    ('spot_count', '必填', '个'),
    ('gene_count', '必填', '个'),
    ('median_genes_per_spot', '必填', '个'),
    ('mean_genes_per_spot', '必填', '个'),
    ('median_counts_per_spot', '必填', '个'),
    ('mean_umi_per_spot', '必填', '个'),
]
spatial_all = spatial_req + [
    ('Unique_Reads', '可选', '个'),
    ('Total_Reads', '可选', '个'),
    ('Valid_Barcode_Reads', '可选', '%'),
    ('Unique_Mapped_Reads', '可选', '%'),
    ('rna_mapping', '可选', '%'),
    ('mean_umi_per_Bin200', '可选', '个'),
    ('Mean Gene Type per Bin200', '可选', '个'),
    ('tissue_coverage', '可选', '%'),
    ('fraction_of_spots_under_tissue', '可选', '%'),
    ('tissue_area_coverage', '可选', '%'),
    ('low_information_spot_rate', '可选', '%'),
    ('image_resolution', '可选', 'px'),
    ('image_contrast_PSNR', '可选', 'dB'),
    ('alignment_quality', '可选', '评分'),
    ('cluster_marker_gene_specificity', '可选', '%'),
    ('tissue_region_gene_expression_consistency', '可选', '%'),
    ('mitochondrial_percentage', '参考', '%'),
    ('ribosomal_percentage', '参考', '%'),
]

atac_req = [
    ('Estimated Number of Cells', '必填', ''),
    ('Median Fragments per Cell', '必填', '个'),
    ('TSS Enrichment Score', '必填', ''),
    ('FRiP', '必填', '%'),
    ('Nucleosome Signal', '必填', ''),
    ('Doublet Rate', '必填', '%'),
    ('Low Quality Cell Rate', '必填', '%'),
]
atac_all = atac_req + [
    ('Capture Efficiency', '可选', '%'),
    ('Cell Cluster Separation', '可选', ''),
    ('Batch Effect Removal R²', '可选', ''),
    ('Cell Type Annotation Rate', '可选', '%'),
    ('Cell Type Consistency', '可选', '%'),
]

# ============================================================
# Create data sheets
# ============================================================
create_data_sheet(wb, 'scRNA-seq填写区', scrna_req + scrna_refs + scrna_opt)
create_data_sheet(wb, 'Spatial填写区', spatial_all)
create_data_sheet(wb, 'scATAC-seq填写区', atac_all)

# ============================================================
# Sheet: 评级汇总
# ============================================================
ws_sum = wb.create_sheet('评级汇总')

ws_sum.merge_cells('A1:G1')
ws_sum['A1'] = '单文件评级汇总表'
ws_sum['A1'].font = TITLE_FONT

r = 3
ws_sum.merge_cells(f'A{r}:G{r}')
ws_sum[f'A{r}'] = '规则：基础等级 = 必填指标中最低等级（木桶效应）；全部可选达1级可升1级'
ws_sum[f'A{r}'].font = NOTE_FONT

r = 5
sum_headers = ['样本名', '数据类型', '必填最低等级', '可选加分', '综合评级', '核心达标/总数', '备注']
for i, h in enumerate(sum_headers, 1):
    ws_sum.cell(row=r, column=i, value=h)
hdr(ws_sum, r, 7)
r += 1

for i in range(20):
    for c in range(1, 8):
        fill = INPUT_FILL if c in [1, 3, 5] else None
        dc(ws_sum, r, c, '', NORMAL, fill, CENTER if c != 7 else WRAP)
    r += 1

# Rating reference
r += 1
ws_sum.merge_cells(f'A{r}:G{r}')
ws_sum[f'A{r}'] = '▼ 综合评级含义'
ws_sum[f'A{r}'].font = SECTION_FONT
r += 1

grades = [
    ['1级', '优质', '所有必填指标达到1级或2级，且可选指标表现良好', '可直接用于高水平研究或AI训练'],
    ['2级', '合格', '所有必填指标在2级及以上', '可用于常规分析'],
    ['3级', '让步接受', '部分必填指标为3级（不超过3项）', '数据可用性受限，需谨慎使用'],
    ['4级', '不合格', '多个必填指标为3级（超过3项）或存在4级', '不建议使用，建议重做或补充'],
]
g_headers = ['等级', '含义', '判定条件', '业务解释']
for i, h in enumerate(g_headers, 1):
    ws_sum.cell(row=r, column=i, value=h)
hdr(ws_sum, r, 4)
r += 1

fills = {1: LEVEL1_FILL, 2: LEVEL2_FILL, 3: LEVEL3_FILL, 4: PatternFill('solid', fgColor='FF4444')}
fonts = {4: Font(name='Microsoft YaHei', bold=True, size=10, color='FFFFFF')}
for row_data in grades:
    grade = int(row_data[0][0])
    for c, val in enumerate(row_data, 1):
        f = fills.get(grade, None)
        fn = fonts.get(grade, NORMAL)
        dc(ws_sum, r, c, val, fn, f, CENTER if c == 1 else WRAP)
    r += 1

ws_sum.column_dimensions['A'].width = 14
ws_sum.column_dimensions['B'].width = 18
ws_sum.column_dimensions['C'].width = 28
ws_sum.column_dimensions['D'].width = 18
ws_sum.column_dimensions['E'].width = 14
ws_sum.column_dimensions['F'].width = 18
ws_sum.column_dimensions['G'].width = 28

# Save
output_path = r"D:\OneDrive - BGI Tech Solutions (Hongkong) Co., Ltd\个人文件\项目\2026\数据整理\cell-standard-docs\docs\qc\数据质控填写表.xlsx"
wb.save(output_path)
print(f"OK: {output_path}")
