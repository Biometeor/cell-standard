from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

header_font = Font(name='Microsoft YaHei', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='4472C4')
normal_font = Font(name='Microsoft YaHei', size=10)
bold_font = Font(name='Microsoft YaHei', bold=True, size=10)
required_fill = PatternFill('solid', fgColor='FFF2CC')
input_fill = PatternFill('solid', fgColor='E8F5E9')
example_fill = PatternFill('solid', fgColor='F5F5F5')
title_font = Font(name='Microsoft YaHei', bold=True, size=14, color='1F4E79')
section_font = Font(name='Microsoft YaHei', bold=True, size=12, color='2E75B6')
thin_border = Border(
    left=Side(style='thin', color='BFBFBF'),
    right=Side(style='thin', color='BFBFBF'),
    top=Side(style='thin', color='BFBFBF'),
    bottom=Side(style='thin', color='BFBFBF')
)
wrap_align = Alignment(wrap_text=True, vertical='center')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

def style_header_row(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

def style_data_cell(ws, row, col, font=None, fill=None):
    cell = ws.cell(row=row, column=col)
    cell.font = font or normal_font
    cell.alignment = wrap_align
    cell.border = thin_border
    if fill:
        cell.fill = fill

# ========== Sheet 1: 数据集元信息 ==========
ws1 = wb.active
ws1.title = '数据集元信息'

ws1.merge_cells('A1:E1')
ws1['A1'] = '数据集元信息（每个项目仅填一份）'
ws1['A1'].font = title_font

headers = ['字段名', '类型', '必填', '描述与参考', '填写值']
for i, h in enumerate(headers, 1):
    ws1.cell(row=3, column=i, value=h)
style_header_row(ws1, 3, 5)

dataset_fields = [
    ['title', 'string', '是', '数据集标题，简要描述。格式：<技术>_<组织/疾病>_<物种>', ''],
    ['summary', 'string', '是', '数据集详细描述，说明实验目的、样本来源、技术平台等（2-5句话）', ''],
    ['contributors', 'string', '是', '递交者姓名，多个用 ; 分隔。例：San Zhang; Si Li', ''],
    ['reference', 'string', '推荐', '发表文献 DOI。例：doi:10.1093/nar/gkad933', ''],
]

for r, row_data in enumerate(dataset_fields, 4):
    for c, val in enumerate(row_data, 1):
        ws1.cell(row=r, column=c, value=val)
        fill = required_fill if c == 3 and val == '是' else None
        font = bold_font if c == 1 else normal_font
        style_data_cell(ws1, r, c, font, fill)
    style_data_cell(ws1, r, 5, normal_font, input_fill)

ws1.column_dimensions['A'].width = 18
ws1.column_dimensions['B'].width = 10
ws1.column_dimensions['C'].width = 8
ws1.column_dimensions['D'].width = 60
ws1.column_dimensions['E'].width = 30

# ========== Sheet 2: 文件元信息-通用字段 ==========
ws2 = wb.create_sheet('文件元信息-通用')

ws2.merge_cells('A1:F1')
ws2['A1'] = '文件元信息（每个样本填一行）—— 通用字段（所有数据类型）'
ws2['A1'].font = title_font

headers2 = ['字段名', '类型', '必填', '参考/可选值', '填写值', '备注']
for i, h in enumerate(headers2, 1):
    ws2.cell(row=3, column=i, value=h)
style_header_row(ws2, 3, 6)

file_fields = [
    ['sample_name', 'string', '是', '样本唯一标识符，建议与文献一致。例：PRJ001', '', '主键，不可重复'],
    ['donor_name', 'string', '是', '供体名称。cell line填na；多供体填pooled；不明填unknown', '', ''],
    ['sequenced_fragment', 'string', '是', '3 prime tag / 5 prime tag / probe-based / full length / not applicable', '', '选其一'],
    ['library_strategy', 'string', '是', 'scRNA-seq / snRNA-seq / scATAC-seq / spatial-transcriptomics', '', '决定处理流程'],
    ['library_construction_method', 'string', '是', "10x 3' v3 / 10x 5' v2 / Smart-seq2 / Drop-seq / 10x Visium / Stereo-seq", '', ''],
    ['development_stage', 'string', '是', '发育阶段。cell line填na；不明填unknown。见参考链接', '', ''],
    ['development_stage_ontology_term_id', 'string', '是', '本体论ID。例：HsapDv:0000016（人成年）', '', ''],
    ['tissue', 'string', '是', '组织名称（自然语言）。例：brain / lung / liver', '', ''],
    ['tissue_ontology_term_id', 'string', '是', 'UBERON ID。cell line用Cellosaurus。例：UBERON:0000955', '', ''],
    ['tissue_type', 'string', '是', 'tissue / organoid / cell line / primary cell culture', '', '选其一'],
    ['sex', 'string', '是', 'male / female / hermaphrodite / unknown', '', '选其一'],
    ['disease', 'string', '是', '疾病状态。健康填 Normal。例：Lung Adenocarcinoma', '', ''],
    ['disease_ontology_term_id', 'string', '是', 'MONDO ID。正常填 MONDO:0000001。例：MONDO:0005069', '', ''],
    ['organism_taxid', 'int', '是', '9606(人) / 10090(小鼠) / 10116(大鼠) / 7955(斑马鱼)', '', ''],
    ['reference_genome', 'string', '是', 'GRCh38 / GRCm39 / mm10 / hg19', '', ''],
    ['gene_annotation_version', 'string', '推荐', 'v110 / GCF_000001405.40 / Ensembl 110', '', ''],
    ['raw_matrix_file_name', 'string', '是', '原始矩阵文件路径。例：./data/S01_matrix.mtx.gz', '', '最原始文件'],
    ['raw_matrix_file_type', 'string', '是', 'h5 / gef / h5ad / mtx / fragments.tsv.gz', '', ''],
    ['raw_matrix_file_md5', 'string', '是', '原始文件MD5校验码（32位hex）', '', ''],
    ['processed_file_name', 'string', '推荐', '分析后H5AD文件路径（可选）。例：./data/S01_analyzed.h5ad', '', ''],
    ['processed_file_md5', 'string', '推荐', '分析后文件MD5校验码', '', ''],
    ['obs_cell_type_column', 'string', '推荐', 'H5AD .obs 中细胞类型列名。例：cell_type', '', ''],
    ['obsm_embedding_key', 'string', '推荐', 'H5AD .obsm 中降维键名。例：X_umap', '', ''],
]

for r, row_data in enumerate(file_fields, 4):
    for c, val in enumerate(row_data, 1):
        ws2.cell(row=r, column=c, value=val)
        fill = required_fill if c == 3 and val == '是' else None
        font = bold_font if c == 1 else normal_font
        style_data_cell(ws2, r, c, font, fill)
    style_data_cell(ws2, r, 5, normal_font, input_fill)

ws2.column_dimensions['A'].width = 28
ws2.column_dimensions['B'].width = 10
ws2.column_dimensions['C'].width = 8
ws2.column_dimensions['D'].width = 55
ws2.column_dimensions['E'].width = 30
ws2.column_dimensions['F'].width = 18

# Dropdowns
dv_lib = DataValidation(type='list', formula1='"scRNA-seq,snRNA-seq,scATAC-seq,spatial-transcriptomics"', allow_blank=True)
dv_lib.error = '请选择有效的文库策略'
ws2.add_data_validation(dv_lib)
dv_lib.add('D7')

dv_frag = DataValidation(type='list', formula1='"3 prime tag,5 prime tag,probe-based,full length,not applicable"', allow_blank=True)
ws2.add_data_validation(dv_frag)
dv_frag.add('D6')

dv_tt = DataValidation(type='list', formula1='"tissue,organoid,cell line,primary cell culture"', allow_blank=True)
ws2.add_data_validation(dv_tt)
dv_tt.add('D11')

dv_sex = DataValidation(type='list', formula1='"male,female,hermaphrodite,unknown"', allow_blank=True)
ws2.add_data_validation(dv_sex)
dv_sex.add('D12')

# ========== Sheet 3: 空间转录组适配字段 ==========
ws3 = wb.create_sheet('空间转录组适配字段')

ws3.merge_cells('A1:E1')
ws3['A1'] = '空间转录组适配字段（仅 Spatial 类型填写）'
ws3['A1'].font = title_font

headers3 = ['字段名', '类型', '必填', '描述', '填写值']
for i, h in enumerate(headers3, 1):
    ws3.cell(row=3, column=i, value=h)
style_header_row(ws3, 3, 5)

spatial_fields = [
    ['image_file', 'string', '条件推荐', '组织图像文件路径（tif/png/jpg）。例：data/S001/spatial/image.tif', ''],
    ['image_file_md5', 'string', '条件推荐', '图像文件MD5校验值', ''],
    ['cell_bin_file', 'string', '条件推荐', '细胞分割/bin结果文件路径。例：data/S001/cell_bin.json', ''],
    ['cell_bin_file_md5', 'string', '条件推荐', '细胞bin文件MD5校验值', ''],
]

for r, row_data in enumerate(spatial_fields, 4):
    for c, val in enumerate(row_data, 1):
        ws3.cell(row=r, column=c, value=val)
        style_data_cell(ws3, r, c, bold_font if c == 1 else normal_font)
    style_data_cell(ws3, r, 5, normal_font, input_fill)

ws3.column_dimensions['A'].width = 22
ws3.column_dimensions['B'].width = 10
ws3.column_dimensions['C'].width = 12
ws3.column_dimensions['D'].width = 55
ws3.column_dimensions['E'].width = 30

# ========== Sheet 4: 补充文件 ==========
ws4 = wb.create_sheet('补充文件')

ws4.merge_cells('A1:E1')
ws4['A1'] = '补充文件（有额外文件需递交时逐行填写）'
ws4['A1'].font = title_font

headers4 = ['字段名', '类型', '必填', '描述与示例', '填写值']
for i, h in enumerate(headers4, 1):
    ws4.cell(row=3, column=i, value=h)
style_header_row(ws4, 3, 5)

supp_fields = [
    ['supplementary_file_name', 'string', '是', '文件名（含后缀），建议含样本ID。例：S01_annotation_script.R', ''],
    ['supplementary_file_type', 'enum', '是', 'Analysis Script / Reference Genome / Image / Metadata / Other', ''],
    ['supplementary_file_description', 'string', '是', '内容用途说明，含软件版本。例：Seurat v4.0 cell type annotation script', ''],
    ['supplementary_file_md5', 'string', '是', '文件MD5校验码', ''],
]

for r, row_data in enumerate(supp_fields, 4):
    for c, val in enumerate(row_data, 1):
        ws4.cell(row=r, column=c, value=val)
        fill = required_fill if c == 3 and val == '是' else None
        style_data_cell(ws4, r, c, bold_font if c == 1 else normal_font, fill)
    style_data_cell(ws4, r, 5, normal_font, input_fill)

ws4.column_dimensions['A'].width = 28
ws4.column_dimensions['B'].width = 10
ws4.column_dimensions['C'].width = 8
ws4.column_dimensions['D'].width = 55
ws4.column_dimensions['E'].width = 30

# ========== Sheet 5: 填写示例 ==========
ws5 = wb.create_sheet('填写示例')

ws5.merge_cells('A1:F1')
ws5['A1'] = '填写示例 —— 仅供参考'
ws5['A1'].font = title_font

ws5.merge_cells('A3:F3')
ws5['A3'] = '示例 1: scRNA-seq 样本'
ws5['A3'].font = section_font

headers5 = ['字段名', '类型', '必填', '参考/可选值', '示例值', '说明']
for i, h in enumerate(headers5, 1):
    ws5.cell(row=4, column=i, value=h)
style_header_row(ws5, 4, 6)

example_data = [
    ['sample_name', 'string', '是', '唯一标识符', 'PRJ001', ''],
    ['donor_name', 'string', '是', '供体名称', 'Patient_001', ''],
    ['sequenced_fragment', 'string', '是', '片段类型', '3 prime tag', ''],
    ['library_strategy', 'string', '是', '实验类型', 'scRNA-seq', ''],
    ['library_construction_method', 'string', '是', '建库方法', "10x 3' v3", ''],
    ['development_stage', 'string', '是', '发育阶段', 'adult', ''],
    ['development_stage_ontology_term_id', 'string', '是', '发育本体ID', 'HsapDv:0000016', ''],
    ['tissue', 'string', '是', '组织名称', 'lung', ''],
    ['tissue_ontology_term_id', 'string', '是', '组织本体ID', 'UBERON:0002048', 'lung'],
    ['tissue_type', 'string', '是', '组织类型', 'tissue', ''],
    ['sex', 'string', '是', '性别', 'male', ''],
    ['disease', 'string', '是', '疾病', 'Lung Adenocarcinoma', ''],
    ['disease_ontology_term_id', 'string', '是', '疾病本体ID', 'MONDO:0005069', ''],
    ['organism_taxid', 'int', '是', '物种TaxID', '9606', '人'],
    ['reference_genome', 'string', '是', '参考基因组', 'GRCh38', ''],
    ['gene_annotation_version', 'string', '推荐', '注释版本', 'Ensembl v110', ''],
    ['raw_matrix_file_name', 'string', '是', '原始矩阵路径', './data/PRJ001_matrix.mtx.gz', ''],
    ['raw_matrix_file_type', 'string', '是', '文件类型', 'mtx', ''],
    ['raw_matrix_file_md5', 'string', '是', 'MD5', 'a1b2c3d4e5f6g7h8i9j0k1l2m3n4o5p6', '32位hex'],
    ['obs_cell_type_column', 'string', '推荐', '细胞类型列名', 'cell_type', ''],
    ['obsm_embedding_key', 'string', '推荐', '降维键名', 'X_umap', ''],
]

for r, row_data in enumerate(example_data, 5):
    for c, val in enumerate(row_data, 1):
        ws5.cell(row=r, column=c, value=val)
        style_data_cell(ws5, r, c, normal_font, example_fill)

ws5.column_dimensions['A'].width = 28
ws5.column_dimensions['B'].width = 10
ws5.column_dimensions['C'].width = 8
ws5.column_dimensions['D'].width = 18
ws5.column_dimensions['E'].width = 35
ws5.column_dimensions['F'].width = 15

# Example 2: Spatial
r_start = 5 + len(example_data) + 2
ws5.merge_cells(f'A{r_start}:F{r_start}')
ws5[f'A{r_start}'] = '示例 2: Spatial 样本（Stereo-seq）'
ws5[f'A{r_start}'].font = section_font

r_start += 1
for i, h in enumerate(headers5, 1):
    ws5.cell(row=r_start, column=i, value=h)
style_header_row(ws5, r_start, 6)

spatial_example = [
    ['sample_name', 'string', '是', '唯一标识符', 'MOB_S001', ''],
    ['library_strategy', 'string', '是', '实验类型', 'spatial-transcriptomics', ''],
    ['library_construction_method', 'string', '是', '建库方法', 'Stereo-seq', ''],
    ['raw_matrix_file_name', 'string', '是', '原始矩阵路径', './data/MOB_S001.bin1.gef', 'bin1 gef'],
    ['raw_matrix_file_type', 'string', '是', '文件类型', 'gef', ''],
    ['image_file', 'string', '条件推荐', '图像文件', './data/MOB_S001/spatial/image.tif', '空间特有'],
    ['cell_bin_file', 'string', '条件推荐', '细胞分割文件', './data/MOB_S001/cell_bin.json', '空间特有'],
]

for r, row_data in enumerate(spatial_example, r_start+1):
    for c, val in enumerate(row_data, 1):
        ws5.cell(row=r, column=c, value=val)
        style_data_cell(ws5, r, c, normal_font, example_fill)

# Save
output_path = r"D:\OneDrive - BGI Tech Solutions (Hongkong) Co., Ltd\个人文件\项目\2026\数据整理\cell-standard-docs\docs\metadata\元信息填写表.xlsx"
wb.save(output_path)
print(f"OK: {output_path}")
