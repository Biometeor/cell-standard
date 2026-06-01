from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# Common styles
HDR_FILL = PatternFill('solid', fgColor='4472C4')
HDR_FONT = Font(name='Microsoft YaHei', bold=True, size=10, color='FFFFFF')
SUB_FILL = PatternFill('solid', fgColor='D6E4F0')
SUB_FONT = Font(name='Microsoft YaHei', bold=True, size=10, color='1F4E79')
TITLE_FONT = Font(name='Microsoft YaHei', bold=True, size=14, color='1F4E79')
SEC_FONT = Font(name='Microsoft YaHei', bold=True, size=12, color='2E75B6')
NORMAL = Font(name='Microsoft YaHei', size=10)
BOLD = Font(name='Microsoft YaHei', bold=True, size=10)
NOTE_FONT = Font(name='Microsoft YaHei', size=9, italic=True, color='666666')
INPUT_FILL = PatternFill('solid', fgColor='F0F8FF')
RESULT_FILL = PatternFill('solid', fgColor='E8F5E9')
L1_FILL = PatternFill('solid', fgColor='C6EFCE')
L2_FILL = PatternFill('solid', fgColor='FFEB9C')
L3_FILL = PatternFill('solid', fgColor='FFC7CE')
L4_FILL = PatternFill('solid', fgColor='FF4444')
YELLOW_FILL = PatternFill('solid', fgColor='FFF2CC')
BORDER = Border(
    left=Side(style='thin', color='BFBFBF'),
    right=Side(style='thin', color='BFBFBF'),
    top=Side(style='thin', color='BFBFBF'),
    bottom=Side(style='thin', color='BFBFBF')
)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
WRAP = Alignment(wrap_text=True, vertical='center')

def hdr(ws, row, ncols):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER
        cell.border = BORDER

def dc(ws, row, col, val, font=None, fill=None, align=None):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = font or NORMAL
    cell.alignment = align or WRAP
    cell.border = BORDER
    if fill:
        cell.fill = fill

# ============================================================
# Sheet 1: 项目基本信息
# ============================================================
ws1 = wb.active
ws1.title = '项目基本信息'

ws1.merge_cells('A1:D1')
ws1['A1'] = '总体数据集自评评分表'
ws1['A1'].font = TITLE_FONT

ws1.merge_cells('A2:D2')
ws1['A2'] = '填写所有 Sheet 后，综合评级自动计算'
ws1['A2'].font = NOTE_FONT

# Project info fields
fields = [
    ('项目名称', '', '项目唯一标识'),
    ('数据类型', 'scRNA-seq / snRNA-seq / scATAC-seq / spatial-transcriptomics', '可多选，用逗号分隔'),
    ('总样本数', '', '递交的样本文件总数'),
    ('总有效细胞/Spot数', '', '所有1-3级文件的有效细胞/Spot数之和'),
    ('项目负责人', '', ''),
    ('递交日期', '', 'YYYY-MM-DD'),
]

r = 4
for i, (name, hint, desc) in enumerate(fields):
    dc(ws1, r+i, 1, name, BOLD, SUB_FILL)
    dc(ws1, r+i, 2, '', NORMAL, INPUT_FILL)
    dc(ws1, r+i, 3, hint, NORMAL, YELLOW_FILL)
    dc(ws1, r+i, 4, desc, NOTE_FONT)

# Project-level statistics (from the standard)
r += len(fields) + 2
ws1.merge_cells(f'A{r}:D{r}')
ws1[f'A{r}'] = '▼ 项目级统计（选填，用于更精确的评估）'
ws1[f'A{r}'].font = SEC_FONT
r += 1

stat_fields = [
    ('project_total_cells/spots', '个', '项目总有效细胞/Spot数（1-3级文件求和）'),
    ('project_avg_score', '分', '单文件QC分数算术平均值'),
    ('qc_consistency (SD)', '', '单文件得分标准差，越小一致性越高'),
    ('project_meta_integrity', '%', '元信息完整度均值'),
]

stat_hdrs = ['字段名', '单位', '说明', '填写值']
for i, h in enumerate(stat_hdrs, 1):
    dc(ws1, r, i, h, HDR_FONT, HDR_FILL, CENTER)
r += 1

for name, unit, desc in stat_fields:
    dc(ws1, r, 1, name, BOLD)
    dc(ws1, r, 2, unit, NORMAL, None, CENTER)
    dc(ws1, r, 3, desc)
    dc(ws1, r, 4, '', NORMAL, INPUT_FILL)
    r += 1

# Spatial-specific stats
r += 1
ws1.merge_cells(f'A{r}:D{r}')
ws1[f'A{r}'] = '空间转录组特有统计（Spatial 项目填写）'
ws1[f'A{r}'].font = SEC_FONT
r += 1

spatial_stat = [
    ('avg_spatial_resolution', 'μm/px', '平均空间分辨率'),
    ('avg_tissue_coverage', '%', '平均组织覆盖率'),
    ('avg_spot_uniformity', '', '平均Spot均匀性（1-5分）'),
    ('avg_mean_genes_per_spot', '个', '平均每Spot基因数'),
    ('avg_mean_umi_per_spot', '个', '平均每Spot UMI数'),
]

for i, h in enumerate(stat_hdrs, 1):
    dc(ws1, r, i, h, HDR_FONT, HDR_FILL, CENTER)
r += 1

for name, unit, desc in spatial_stat:
    dc(ws1, r, 1, name, BOLD)
    dc(ws1, r, 2, unit, NORMAL, None, CENTER)
    dc(ws1, r, 3, desc)
    dc(ws1, r, 4, '', NORMAL, INPUT_FILL)
    r += 1

ws1.column_dimensions['A'].width = 32
ws1.column_dimensions['B'].width = 18
ws1.column_dimensions['C'].width = 40
ws1.column_dimensions['D'].width = 28

# ============================================================
# Sheet 2: 元信息完整度自评
# ============================================================
ws2 = wb.create_sheet('元信息完整度自评')

ws2.merge_cells('A1:D1')
ws2['A1'] = '元信息完整度自评 — 检查每个字段是否已填报'
ws2['A1'].font = TITLE_FONT
ws2['A2'] = '在"已填报"列标记 ✓ 或 ✗，系统自动计算完整度'
ws2['A2'].font = NOTE_FONT

r = 4
# Core fields
ws2.merge_cells(f'A{r}:D{r}')
ws2[f'A{r}'] = '◆ 核心字段（5项 — 决定数据是否有基本生物学价值）'
ws2[f'A{r}'].font = SEC_FONT
r += 1

core_headers = ['#', '字段', '说明', '已填报(✓/✗)']
for i, h in enumerate(core_headers, 1):
    dc(ws2, r, i, h, HDR_FONT, HDR_FILL, CENTER)
r += 1

core_fields = [
    ('1', 'organism_taxid', '物种TaxID（9606/10090等）'),
    ('2', 'tissue / tissue_ontology_term_id', '组织名称及UBERON ID'),
    ('3', 'disease / disease_ontology_term_id', '疾病状态及MONDO ID'),
    ('4', 'sex', '性别（male/female/unknown）'),
    ('5', 'development_stage / ontology_id', '发育阶段及本体论ID'),
]

for i, (num, field, desc) in enumerate(core_fields):
    dc(ws2, r, 1, num, NORMAL, None, CENTER)
    dc(ws2, r, 2, field, BOLD)
    dc(ws2, r, 3, desc)
    dc(ws2, r, 4, '', NORMAL, INPUT_FILL, CENTER)
    r += 1

# Core summary
core_summary_row = r
dc(ws2, r, 1, '', NORMAL, SUB_FILL)
dc(ws2, r, 2, '核心字段已填数', BOLD, SUB_FILL)
dc(ws2, r, 3, '= 已填 ✓ 的个数（自动统计）', NORMAL, YELLOW_FILL)
dc(ws2, r, 4, '=COUNTIF(D5:D9,"✓")', NORMAL, RESULT_FILL, CENTER)
r += 1

dc(ws2, r, 1, '', NORMAL, SUB_FILL)
dc(ws2, r, 2, '核心完整度', BOLD, SUB_FILL)
dc(ws2, r, 3, '= 已填数 ÷ 5 × 100%', NORMAL, YELLOW_FILL)
dc(ws2, r, 4, '=D10/5', NORMAL, RESULT_FILL, CENTER)
ws2.cell(row=r, column=4).number_format = '0%'
r += 1

# Recommended fields
r += 1
ws2.merge_cells(f'A{r}:D{r}')
ws2[f'A{r}'] = '◆ 推荐字段（16项 — 决定数据是否"好用"）'
ws2[f'A{r}'].font = SEC_FONT
r += 1

for i, h in enumerate(core_headers, 1):
    dc(ws2, r, i, h, HDR_FONT, HDR_FILL, CENTER)
r += 1

rec_fields = [
    # 管理属性
    ('', '【管理属性】', '', ''),
    ('', 'title', '数据集标题', ''),
    ('', 'summary', '数据集详细描述', ''),
    ('', 'contributors', '递交者姓名', ''),
    ('', 'reference', '文献DOI号', ''),
    # 样本属性
    ('', '【样本属性】', '', ''),
    ('', 'sample_name', '样本唯一标识符', ''),
    ('', 'donor_name', '供体/个体名称', ''),
    # 实验属性
    ('', '【实验属性】', '', ''),
    ('', 'library_strategy', '实验类型（scRNA-seq等）', ''),
    ('', 'library_construction_method', '文库构建方法', ''),
    ('', 'sequenced_fragment', '测序片段类型', ''),
    ('', 'tissue_type', '组织类型（tissue/organoid等）', ''),
    # 生信属性
    ('', '【生信属性】', '', ''),
    ('', 'reference_genome', '参考基因组版本', ''),
    ('', 'gene_annotation_version', '基因注释版本', ''),
    # 文件与结构属性
    ('', '【文件与结构属性】', '', ''),
    ('', 'raw_matrix_file', '原始矩阵文件', ''),
    ('', 'processed_file', '分析后文件（可选）', ''),
    ('', 'obs_cell_type_column', '细胞类型注释列名', ''),
    ('', 'obsm_embedding_key', '降维坐标键名', ''),
]

for _ in rec_fields:
    num, field, desc, _ = _
    if field.startswith('【'):
        ws2.merge_cells(f'A{r}:D{r}')
        dc(ws2, r, 1, field, SUB_FONT, SUB_FILL)
    else:
        dc(ws2, r, 1, num, NORMAL, None, CENTER)
        dc(ws2, r, 2, field, BOLD)
        dc(ws2, r, 3, desc)
        dc(ws2, r, 4, '', NORMAL, INPUT_FILL, CENTER)
    r += 1

# Recommended summary
rec_summary_row = r
dc(ws2, r, 1, '', NORMAL, SUB_FILL)
dc(ws2, r, 2, '推荐字段已填数', BOLD, SUB_FILL)
dc(ws2, r, 3, '= 已填 ✓ 的个数（不含分组标题行）', NORMAL, YELLOW_FILL)
dc(ws2, r, 4, f'=COUNTA(D{rec_summary_row-16}:D{r-1})', NORMAL, RESULT_FILL, CENTER)
r += 1

dc(ws2, r, 1, '', NORMAL, SUB_FILL)
dc(ws2, r, 2, '辅助完整度', BOLD, SUB_FILL)
dc(ws2, r, 3, '= 已填数 ÷ 16 × 100%', NORMAL, YELLOW_FILL)
dc(ws2, r, 4, f'=D{rec_summary_row}/16', NORMAL, RESULT_FILL, CENTER)
ws2.cell(row=r, column=4).number_format = '0%'
r += 1

# Meta rating result
r += 1
ws2.merge_cells(f'A{r}:D{r}')
ws2[f'A{r}'] = '▼ 元信息评级结果'
ws2[f'A{r}'].font = SEC_FONT
r += 1

core_pct_cell = f'D{core_summary_row+1}'  # core completeness
aux_pct_cell = f'D{r-3}'  # aux completeness (row before last)

meta_rules = [
    ('核心完整度', f'={core_pct_cell}', '', ''),
    ('辅助完整度', f'={aux_pct_cell}', '', ''),
    ('元信息评级', '', '', '优/良/差'),
]

for i, h in enumerate(['指标', '数值', '', '结果'], 1):
    dc(ws2, r, i, h, HDR_FONT, HDR_FILL, CENTER)
r += 1

dc(ws2, r, 1, '核心完整度', BOLD)
dc(ws2, r, 2, f'=D{core_summary_row+1}', NORMAL, RESULT_FILL, CENTER)
ws2.cell(row=r, column=2).number_format = '0%'
dc(ws2, r, 3, '≥50% 通过', NOTE_FONT)
dc(ws2, r, 4, '', NORMAL, INPUT_FILL, CENTER)
r += 1

dc(ws2, r, 1, '辅助完整度', BOLD)
dc(ws2, r, 2, f'=D{rec_summary_row+1}', NORMAL, RESULT_FILL, CENTER)
ws2.cell(row=r, column=2).number_format = '0%'
dc(ws2, r, 3, '≥70% 为优', NOTE_FONT)
dc(ws2, r, 4, '', NORMAL, INPUT_FILL, CENTER)
r += 1

dc(ws2, r, 1, '元信息评级', BOLD)
dc(ws2, r, 2, '', NORMAL, RESULT_FILL, CENTER)
dc(ws2, r, 3, '优:核心=100%且辅助≥70%；良:核心≥50%；差:核心<50%', NOTE_FONT)
dc(ws2, r, 4, '', NORMAL, INPUT_FILL, CENTER)
r += 1

ws2.column_dimensions['A'].width = 10
ws2.column_dimensions['B'].width = 36
ws2.column_dimensions['C'].width = 40
ws2.column_dimensions['D'].width = 18

# ============================================================
# Sheet 3: 单文件QC分布
# ============================================================
ws3 = wb.create_sheet('单文件QC分布')

ws3.merge_cells('A1:D1')
ws3['A1'] = '单文件 QC 评级分布 — 填写各等级文件数量'
ws3['A1'].font = TITLE_FONT
ws3['A2'] = '根据数据质控填写表的评级结果，统计每个等级有多少个文件'
ws3['A2'].font = NOTE_FONT

r = 4
headers = ['等级', '定义', '文件数量', '占比']
for i, h in enumerate(headers, 1):
    dc(ws3, r, i, h, HDR_FONT, HDR_FILL, CENTER)
r += 1

level_info = [
    ('1级', '优质', L1_FILL),
    ('2级', '合格', L2_FILL),
    ('3级', '让步接受', L3_FILL),
    ('4级', '不合格', L4_FILL),
]

total_row = r + 4  # row for total
for level, desc, fill in level_info:
    dc(ws3, r, 1, level, BOLD, fill, CENTER)
    dc(ws3, r, 2, desc, NORMAL)
    dc(ws3, r, 3, 0, NORMAL, INPUT_FILL, CENTER)
    dc(ws3, r, 4, f'=C{r}/C{total_row}', NORMAL, RESULT_FILL, CENTER)
    ws3.cell(row=r, column=4).number_format = '0.0%'
    r += 1

# Total
dc(ws3, r, 1, '合计', BOLD, SUB_FILL, CENTER)
dc(ws3, r, 2, '', NORMAL, SUB_FILL)
dc(ws3, r, 3, '=SUM(C4:C7)', BOLD, RESULT_FILL, CENTER)
dc(ws3, r, 4, '=SUM(D4:D7)', NORMAL, RESULT_FILL, CENTER)
ws3.cell(row=r, column=4).number_format = '0.0%'

# Base quality tier
r += 2
ws3.merge_cells(f'A{r}:D{r}')
ws3[f'A{r}'] = '▼ 基础质量定级（自动计算）'
ws3[f'A{r}'].font = SEC_FONT
r += 1

tier_headers = ['条件', '等级', '判定', '结果']
for i, h in enumerate(tier_headers, 1):
    dc(ws3, r, i, h, HDR_FONT, HDR_FILL, CENTER)
r += 1

tier_rules = [
    ('A类(优质基线)', '80%以上文件为1级，且总细胞/Spot数>50000', 'A'),
    ('B类(合格基线)', '80%以上文件≥2级，无4级文件', 'B'),
    ('C类(风险基线)', '存在4级文件，或20%以上文件为3级', 'C'),
]

for cond, rule, tier_label in tier_rules:
    dc(ws3, r, 1, cond, BOLD)
    dc(ws3, r, 2, rule, NORMAL, YELLOW_FILL)
    dc(ws3, r, 3, tier_label, NORMAL, None, CENTER)
    dc(ws3, r, 4, '', NORMAL, INPUT_FILL, CENTER)
    r += 1

# Manual tier selection
r += 1
dc(ws3, r, 1, '本项目基础质量等级', BOLD, SUB_FILL)
dc(ws3, r, 2, '参考上方规则，选择 A/B/C', NOTE_FONT)
dc(ws3, r, 3, '', NORMAL, INPUT_FILL, CENTER)
dc(ws3, r, 4, '', NORMAL)

ws3.column_dimensions['A'].width = 26
ws3.column_dimensions['B'].width = 38
ws3.column_dimensions['C'].width = 16
ws3.column_dimensions['D'].width = 16

# ============================================================
# Sheet 4: 综合评级计算
# ============================================================
ws4 = wb.create_sheet('综合评级计算')

ws4.merge_cells('A1:D1')
ws4['A1'] = '综合评级计算 — 填写基础质量和元信息评级后自动得出'
ws4['A1'].font = TITLE_FONT

r = 3
ws4.merge_cells(f'A{r}:D{r}')
ws4[f'A{r}'] = '参考 project_evaluation.md 的评级流程'
ws4[f'A{r}'].font = NOTE_FONT
r += 1

# Input section
inputs = [
    ('基础质量等级', 'A/B/C', '从"单文件QC分布"Sheet 填入'),
    ('元信息评级', '优/良/差', '从"元信息完整度自评"Sheet 填入'),
]

in_headers = ['输入项', '填写值', '参考', '说明']
for i, h in enumerate(in_headers, 1):
    dc(ws4, r, i, h, HDR_FONT, HDR_FILL, CENTER)
r += 1

base_tier_cell = f'C{r}'
meta_rating_cell = f'C{r+1}'

for label, hint, desc in inputs:
    dc(ws4, r, 1, label, BOLD, SUB_FILL)
    dc(ws4, r, 2, '', NORMAL, INPUT_FILL, CENTER)
    dc(ws4, r, 3, hint, NORMAL, YELLOW_FILL, CENTER)
    dc(ws4, r, 4, desc)
    r += 1

# Rating matrix
r += 1
ws4.merge_cells(f'A{r}:D{r}')
ws4[f'A{r}'] = '▼ 评级矩阵（参考）'
ws4[f'A{r}'].font = SEC_FONT
r += 1

matrix = [
    ['基础\\元信息', '优', '良', '差'],
    ['A类', '1级', '2级', '4级（一票否决）'],
    ['B类', '2级', '2级', '4级（一票否决）'],
    ['C类', '2级', '3级', '4级（一票否决）'],
]

for row_data in matrix:
    for c, val in enumerate(row_data, 1):
        if row_data == matrix[0]:
            dc(ws4, r, c, val, HDR_FONT, HDR_FILL, CENTER)
        else:
            f = SUB_FILL if c == 1 else None
            dc(ws4, r, c, val, BOLD if c == 1 else NORMAL, f, CENTER)
    r += 1

# Final result
r += 2
ws4.merge_cells(f'A{r}:D{r}')
ws4[f'A{r}'] = '▼ 最终评级结果'
ws4[f'A{r}'].font = SEC_FONT
r += 1

res_headers = ['维度', '输入/计算结果', '说明', '']
for i, h in enumerate(res_headers, 1):
    dc(ws4, r, i, h, HDR_FONT, HDR_FILL, CENTER)
r += 1

results = [
    ('基础质量等级', '', 'A/B/C', ''),
    ('元信息评级', '', '优/良/差', ''),
    ('综合评级', '', '1级/2级/3级/4级', ''),
    ('判定说明', '', '', '参照评级矩阵交叉得出'),
]

for label, _, hint, desc in results:
    dc(ws4, r, 1, label, BOLD, SUB_FILL)
    dc(ws4, r, 2, '', NORMAL, RESULT_FILL, CENTER)
    dc(ws4, r, 3, hint, NOTE_FONT)
    dc(ws4, r, 4, desc)
    r += 1

ws4.column_dimensions['A'].width = 20
ws4.column_dimensions['B'].width = 20
ws4.column_dimensions['C'].width = 28
ws4.column_dimensions['D'].width = 30

# ============================================================
# Sheet 5: 评级标准参考
# ============================================================
ws5 = wb.create_sheet('评级标准参考')

ws5.merge_cells('A1:D1')
ws5['A1'] = '评级标准参考 — 填写前请阅读'
ws5['A1'].font = TITLE_FONT

# Section 1: Meta-info completeness
r = 3
ws5.merge_cells(f'A{r}:D{r}')
ws5[f'A{r}'] = '▼ 元信息完整度评级'
ws5[f'A{r}'].font = SEC_FONT
r += 1

meta_hdrs = ['评级', '判定条件', '业务解释', '']
for i, h in enumerate(meta_hdrs, 1):
    dc(ws5, r, i, h, HDR_FONT, HDR_FILL, CENTER)
r += 1

meta_grades = [
    ('优', '核心完整度=100% 且 辅助完整度≥70%', '精品数据，可直接用于深度分析或模型训练'),
    ('良', '核心完整度≥50% 但 辅助完整度<70%', '基础可用，适合先收进来再补充'),
    ('差', '核心完整度<50%（5个核心缺3个及以上）', '直接否决，毫无生物学价值，不予归档'),
]

for grade, cond, desc in meta_grades:
    f = L1_FILL if grade == '优' else (L2_FILL if grade == '良' else L3_FILL)
    dc(ws5, r, 1, grade, BOLD, f, CENTER)
    dc(ws5, r, 2, cond)
    dc(ws5, r, 3, desc)
    dc(ws5, r, 4, '', NORMAL)
    r += 1

# Section 2: Base quality tier
r += 1
ws5.merge_cells(f'A{r}:D{r}')
ws5[f'A{r}'] = '▼ 基础质量定级'
ws5[f'A{r}'].font = SEC_FONT
r += 1

for i, h in enumerate(['等级', '判定条件', '', ''], 1):
    dc(ws5, r, i, h, HDR_FONT, HDR_FILL, CENTER)
r += 1

tiers = [
    ('A类（优质基线）', '80%以上单文件为1级，且项目总细胞/Spot数>50000'),
    ('B类（合格基线）', '80%以上单文件为2级及以上，无4级文件'),
    ('C类（风险基线）', '存在4级文件，或20%以上单文件为3级'),
]

for tier, cond in tiers:
    dc(ws5, r, 1, tier, BOLD)
    dc(ws5, r, 2, cond, NORMAL, YELLOW_FILL)
    dc(ws5, r, 3, '', NORMAL)
    dc(ws5, r, 4, '', NORMAL)
    r += 1

# Section 3: Final comprehensive rating
r += 1
ws5.merge_cells(f'A{r}:D{r}')
ws5[f'A{r}'] = '▼ 最终综合评级'
ws5[f'A{r}'].font = SEC_FONT
r += 1

for i, h in enumerate(['最终等级', '定义', '评判标准', ''], 1):
    dc(ws5, r, i, h, HDR_FONT, HDR_FILL, CENTER)
r += 1

final_grades = [
    ('1级', '优质', '基础=A类 且 元信息=优。数据质量极佳且信息完备'),
    ('2级', '合格', '基础=A/B类+元信息=良，或 基础=C类+元信息=优'),
    ('3级', '让步接受', '基础=C类 且 元信息=良。数据可用性受限'),
    ('4级', '不合格', '元信息=差（一票否决），无论基础质量如何'),
]

fills_map = {'1级': L1_FILL, '2级': L2_FILL, '3级': L3_FILL, '4级': L4_FILL}
for grade, name, desc in final_grades:
    f = fills_map.get(grade, None)
    fn = Font(name='Microsoft YaHei', bold=True, size=10, color='FFFFFF' if grade == '4级' else '000000')
    dc(ws5, r, 1, grade, fn, f, CENTER)
    dc(ws5, r, 2, name, BOLD)
    dc(ws5, r, 3, desc)
    dc(ws5, r, 4, '', NORMAL)
    r += 1

ws5.column_dimensions['A'].width = 22
ws5.column_dimensions['B'].width = 42
ws5.column_dimensions['C'].width = 44
ws5.column_dimensions['D'].width = 10

# Save
output_path = r"D:\OneDrive - BGI Tech Solutions (Hongkong) Co., Ltd\个人文件\项目\2026\数据整理\cell-standard-docs\docs\evaluation\数据集自评评分表.xlsx"
wb.save(output_path)
print(f"OK: {output_path}")
